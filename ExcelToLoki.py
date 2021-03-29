from openpyxl import load_workbook

def main():

    hashcount = 0
    c2count = 0
    
    wb = load_workbook('2021年3月25日IOCs.xlsx')
    sheet = wb.active
    
    with open('hash-APT-202103.txt','a',encoding='utf-8') as file_handle:
        file_handle.write('\n\n')
        for row in sheet.rows:
            if row[0].value in ('md5','sha256','sha1'):
                i = 0
                hashcount = hashcount + 1
                for col in row:
                    if col.value in ('md5','sha256','sha1',' '):
                        continue
                    i = i + 1
                    if i==1:
                        file_handle.write(str(col.value))
                        file_handle.write(';')
                    else:
                        file_handle.write(str(col.value))
                        file_handle.write(' ')
                file_handle.write('\n')
    file_handle.close
            
    with open('c2-APT-202103.txt','a',encoding='utf-8') as file_handle:
        file_handle.write('\n\n')
        for row in sheet.rows:
            if row[0].value in ('IP','Domains','URL'):
                i = 0
                c2count = c2count + 1
                for col in row:
                    if col.value in ('IP','Domains','URL',' '):
                        continue
                    i = i + 1 
                    if i==1:
                        file_handle.write(str(col.value))
                file_handle.write('\n')
    file_handle.close  

    print("total hash:",hashcount)
    print("total c2:",c2count)
    if hashcount + c2count + 1 == sheet.max_row:
        print("everything is OK!!")


if __name__=="__main__":
    main()
